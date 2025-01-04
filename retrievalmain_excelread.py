# main_retrieval.py
import os
from dotenv import load_dotenv
from retrieval_utils import LegalRetrievalSystem
from elasticsearch_utils import ElasticsearchManager
from neo4j_manager import Neo4jManager
from models import EmbeddingModel
import pandas as pd
import openpyxl

def main():
    # 載入環境變數
    load_dotenv()

    # 初始化各個組件
    es_manager = ElasticsearchManager(
        host="https://localhost:9200",
        username=os.getenv('ELASTIC_USER'),
        password=os.getenv('ELASTIC_PASSWORD')
    )
    
    neo4j_manager = Neo4jManager(
        uri=os.getenv('NEO4J_URI'),
        user=os.getenv('NEO4J_USER'),
        password=os.getenv('NEO4J_PASSWORD')
    )
    
    embedding_model = EmbeddingModel()

    # 初始化檢索系統
    retrieval_system = LegalRetrievalSystem(
        elasticsearch_manager=es_manager,
        neo4j_manager=neo4j_manager,
        embedding_model=embedding_model
    )

    try:
        # Get input Excel filename
        input_filename = input("Please enter the input Excel filename: ")
        
        # Load Excel file and show available sheets
        wb = openpyxl.load_workbook(input_filename)
        print("\nAvailable sheets:", wb.sheetnames)
        
        # Get sheet name from user
        sheet_name = input("Please enter the sheet name to use: ")
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
        
        # Load sheet and show columns
        df = pd.read_excel(input_filename, sheet_name=sheet_name)
        print("\nAvailable columns:", list(df.columns))
        
        # Get column name from user
        column_name = input("Please enter the column name containing the input data: ")
        if column_name not in df.columns:
            raise ValueError(f"Column '{column_name}' not found in sheet")
        
        # Show row range
        print(f"\nAvailable row range: 0 to {len(df[column_name]) - 1}")
        
        # Get row range from user
        start_row = int(input("Enter start row number (base 0): "))
        end_row = int(input("Enter end row number (base 0): "))
        
        if start_row < 0 or end_row >= len(df[column_name]) or start_row > end_row:
            raise ValueError("Invalid row range")

        # Create output Excel file
        output_filename = "kenneth85-llama-3-taiwan_8b-instruct-dpo.xlsx"
        output_wb = openpyxl.Workbook()
        output_ws = output_wb.active
        output_row = 1  # Start from first row

        # Process each input row
        for idx in range(start_row, end_row + 1):
            input_text = str(df[column_name].iloc[idx])
            print(f"\nProcessing row {idx}...")
            
            try:
                # Process case and get response
                response = retrieval_system.process_case(input_text)
                
                # Write to output Excel
                output_ws.cell(row=output_row, column=1, value=input_text)
                output_ws.cell(row=output_row, column=2, value=response)
                output_row += 1
                
                print(f"Processed row {idx} successfully")
            except Exception as e:
                print(f"Error processing row {idx}: {str(e)}")
                # Still write to output Excel, but mark as error
                output_ws.cell(row=output_row, column=1, value=input_text)
                output_ws.cell(row=output_row, column=2, value=f"ERROR: {str(e)}")
                output_row += 1

        # Save output file
        output_wb.save(output_filename)
        print(f"\nProcessing complete. Results saved to {output_filename}")

    except Exception as e:
        print(f"An error occurred: {str(e)}")
    finally:
        neo4j_manager.close()

if __name__ == "__main__":
    main()